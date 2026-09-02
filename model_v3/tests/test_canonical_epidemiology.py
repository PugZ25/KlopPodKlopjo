from __future__ import annotations

import csv
import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from model_v3.data.canonical_epidemiology import (
    DataValidationError,
    build_calendar,
    combine_weekly_cases,
    load_municipality_dimension,
    load_population_dataset,
    load_weekly_case_workbook,
    write_csv_dataset,
)


FIXTURE_ROOT = Path(__file__).parent / "fixtures"
MUNICIPALITY_CONFIG = {
    "code_property": "SIFRA",
    "name_property": "NAZIV",
}
POPULATION_CONFIG = {
    "measure_dimension": "MERITVE",
    "measure_code": "45",
    "expected_measure_label": "Population - Total - 1 January",
    "municipality_dimension": "OBČINE",
    "year_dimension": "LETO",
}
WEEKLY_CONFIG = {
    "header_row": 1,
    "municipality_name_column": 2,
    "municipality_header": "Občina bivališča\n/Obolenja po tednih\n",
    "total_label": "SKUPAJ",
    "week_header_pattern": r"^(?P<year>\d{4})-(?P<week>\d{2})$",
    "blank_case_cell_value": 0,
    "blank_case_cell_rule_source": "User-confirmed project rule",
    "municipality_name_alias_to_code": {},
}


def write_case_fixture(
    path: Path,
    municipality_rows: list[tuple[str, object, object, object]],
    *,
    aggregate_week_1: object,
    aggregate_week_2: object,
    aggregate_total: object,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "2020"
    worksheet.cell(1, 2, WEEKLY_CONFIG["municipality_header"])
    worksheet.cell(1, 5, "2020-01")
    worksheet.cell(1, 6, "2020-02")
    worksheet.cell(1, 7, "SKUPAJ")
    for row_index, (name, week_1, week_2, total) in enumerate(
        municipality_rows, start=2
    ):
        worksheet.cell(row_index, 2, name)
        worksheet.cell(row_index, 5, week_1)
        worksheet.cell(row_index, 6, week_2)
        worksheet.cell(row_index, 7, total)
    aggregate_row = len(municipality_rows) + 2
    worksheet.cell(aggregate_row, 2, "SKUPAJ")
    worksheet.cell(aggregate_row, 5, aggregate_week_1)
    worksheet.cell(aggregate_row, 6, aggregate_week_2)
    worksheet.cell(aggregate_row, 7, aggregate_total)
    workbook.save(path)
    workbook.close()


class CanonicalEpidemiologyTests(unittest.TestCase):
    def setUp(self) -> None:
        self.municipalities, self.municipality_quality = load_municipality_dimension(
            FIXTURE_ROOT / "municipality_source.geojson",
            MUNICIPALITY_CONFIG,
            code_width=3,
        )

    def test_small_fixtures_convert_case_blanks_to_zero_but_preserve_population_null(
        self,
    ) -> None:
        population, population_quality = load_population_dataset(
            FIXTURE_ROOT / "population_source.json",
            POPULATION_CONFIG,
            self.municipalities,
            code_width=3,
        )
        self.assertEqual([row["municipality_code"] for row in self.municipalities], ["001", "002"])
        self.assertIsNone(
            next(
                row["population"]
                for row in population
                if row["municipality_code"] == "002" and row["year"] == 2020
            )
        )
        self.assertEqual(population_quality["missing_population_count"], 1)
        self.assertEqual(population_quality["explicit_zero_count"], 0)
        self.assertEqual(population_quality["population_codes_not_in_municipality"], [])

        with tempfile.TemporaryDirectory() as temporary_directory:
            fixture_path = Path(temporary_directory) / "weekly.xlsx"
            write_case_fixture(
                fixture_path,
                [
                    ("Alpha", 1, None, 1),
                    ("Beta", 0, 2, 2),
                ],
                aggregate_week_1=1,
                aggregate_week_2=2,
                aggregate_total=3,
            )
            lyme_rows, lyme_quality = load_weekly_case_workbook(
                fixture_path,
                "lyme_cases",
                WEEKLY_CONFIG,
                self.municipalities,
                code_width=3,
            )
            kme_rows, _ = load_weekly_case_workbook(
                fixture_path,
                "kme_cases",
                WEEKLY_CONFIG,
                self.municipalities,
                code_width=3,
            )
            weekly_rows, weekly_quality = combine_weekly_cases(
                {"lyme_cases": lyme_rows, "kme_cases": kme_rows},
                self.municipalities,
            )
            calendar = build_calendar(weekly_rows)

            alpha_week_2 = next(
                row
                for row in weekly_rows
                if row["municipality_code"] == "001"
                and row["issue_week"].isoformat() == "2020-01-06"
            )
            beta_week_1 = next(
                row
                for row in weekly_rows
                if row["municipality_code"] == "002"
                and row["issue_week"].isoformat() == "2019-12-30"
            )
            self.assertEqual(alpha_week_2["lyme_cases"], 0)
            self.assertEqual(beta_week_1["lyme_cases"], 0)
            self.assertEqual(lyme_quality["source_blank_case_cell_count"], 1)
            self.assertEqual(lyme_quality["source_explicit_zero_cell_count"], 1)
            self.assertEqual(lyme_quality["canonical_zero_count"], 2)
            self.assertEqual(lyme_quality["missing_case_value_count"], 0)
            self.assertEqual(
                lyme_quality["blank_case_cell_rule"],
                {
                    "canonical_value": 0,
                    "source": "User-confirmed project rule",
                },
            )
            self.assertEqual(
                lyme_quality["sheets"]["2020"]["missing_week_numbers"],
                list(range(3, 54)),
            )
            self.assertEqual(weekly_quality["missing_municipality_week_keys_count"], 0)
            self.assertEqual(calendar[0]["issue_week"].isoformat(), "2019-12-30")
            self.assertEqual(calendar[0]["year"], 2020)
            self.assertEqual(calendar[0]["iso_week"], 1)

            csv_path = Path(temporary_directory) / "weekly.csv"
            write_csv_dataset(
                csv_path,
                weekly_rows,
                ["municipality_code", "issue_week", "lyme_cases", "kme_cases"],
            )
            with csv_path.open(encoding="utf-8", newline="") as handle:
                csv_rows = list(csv.DictReader(handle))
            csv_alpha_week_2 = next(
                row
                for row in csv_rows
                if row["municipality_code"] == "001"
                and row["issue_week"] == "2020-01-06"
            )
            csv_beta_week_1 = next(
                row
                for row in csv_rows
                if row["municipality_code"] == "002"
                and row["issue_week"] == "2019-12-30"
            )
            self.assertEqual(csv_alpha_week_2["lyme_cases"], "0")
            self.assertEqual(csv_beta_week_1["lyme_cases"], "0")

    def test_duplicate_municipality_code_is_rejected(self) -> None:
        payload = json.loads(
            (FIXTURE_ROOT / "municipality_source.geojson").read_text(encoding="utf-8")
        )
        payload["features"][1]["properties"]["SIFRA"] = 1
        with tempfile.TemporaryDirectory() as temporary_directory:
            path = Path(temporary_directory) / "duplicate.geojson"
            path.write_text(json.dumps(payload), encoding="utf-8")
            with self.assertRaisesRegex(DataValidationError, "duplicate keys"):
                load_municipality_dimension(
                    path,
                    MUNICIPALITY_CONFIG,
                    code_width=3,
                )

    def test_blank_case_rule_must_remain_zero(self) -> None:
        invalid_config = dict(WEEKLY_CONFIG)
        invalid_config["blank_case_cell_value"] = None
        with self.assertRaisesRegex(DataValidationError, "must be configured as 0"):
            load_weekly_case_workbook(
                Path("not-read.xlsx"),
                "lyme_cases",
                invalid_config,
                self.municipalities,
                code_width=3,
            )

    def test_negative_population_is_rejected(self) -> None:
        payload = json.loads(
            (FIXTURE_ROOT / "population_source.json").read_text(encoding="utf-8")
        )
        payload["value"][0] = -1
        with tempfile.TemporaryDirectory() as temporary_directory:
            path = Path(temporary_directory) / "negative_population.json"
            path.write_text(json.dumps(payload), encoding="utf-8")
            with self.assertRaisesRegex(DataValidationError, "must not be negative"):
                load_population_dataset(
                    path,
                    POPULATION_CONFIG,
                    self.municipalities,
                    code_width=3,
                )

    def test_negative_case_count_is_rejected(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            path = Path(temporary_directory) / "negative_cases.xlsx"
            write_case_fixture(
                path,
                [("Alpha", -1, None, None), ("Beta", None, None, None)],
                aggregate_week_1=None,
                aggregate_week_2=None,
                aggregate_total=None,
            )
            with self.assertRaisesRegex(DataValidationError, "must not be negative"):
                load_weekly_case_workbook(
                    path,
                    "lyme_cases",
                    WEEKLY_CONFIG,
                    self.municipalities,
                    code_width=3,
                )

    def test_municipality_row_total_mismatch_is_rejected(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            path = Path(temporary_directory) / "row_total_mismatch.xlsx"
            write_case_fixture(
                path,
                [("Alpha", 1, None, 2), ("Beta", None, None, None)],
                aggregate_week_1=1,
                aggregate_week_2=None,
                aggregate_total=1,
            )
            with self.assertRaisesRegex(
                DataValidationError, "municipality-row total mismatches"
            ):
                load_weekly_case_workbook(
                    path,
                    "lyme_cases",
                    WEEKLY_CONFIG,
                    self.municipalities,
                    code_width=3,
                )

    def test_aggregate_total_mismatch_is_reported_without_changing_cells(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            path = Path(temporary_directory) / "aggregate_mismatch.xlsx"
            write_case_fixture(
                path,
                [("Alpha", 1, None, 1), ("Beta", 0, 2, 2)],
                aggregate_week_1=2,
                aggregate_week_2=2,
                aggregate_total=4,
            )
            rows, quality = load_weekly_case_workbook(
                path,
                "lyme_cases",
                WEEKLY_CONFIG,
                self.municipalities,
                code_width=3,
            )
            self.assertEqual(len(quality["aggregate_total_mismatches"]), 1)
            alpha_week_1 = next(
                row
                for row in rows
                if row["municipality_code"] == "001"
                and row["issue_week"].isoformat() == "2019-12-30"
            )
            self.assertEqual(alpha_week_1["lyme_cases"], 1)

    def test_unmatched_municipality_name_is_rejected_and_named(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            path = Path(temporary_directory) / "unmatched.xlsx"
            write_case_fixture(
                path,
                [("UNKNOWN PLACE", None, None, None)],
                aggregate_week_1=None,
                aggregate_week_2=None,
                aggregate_total=None,
            )
            with self.assertRaisesRegex(DataValidationError, "UNKNOWN PLACE"):
                load_weekly_case_workbook(
                    path,
                    "lyme_cases",
                    WEEKLY_CONFIG,
                    self.municipalities,
                    code_width=3,
                )

    def test_duplicate_weekly_municipality_row_is_rejected(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            path = Path(temporary_directory) / "duplicate_weekly.xlsx"
            write_case_fixture(
                path,
                [("Alpha", None, None, None), ("Alpha", None, None, None)],
                aggregate_week_1=None,
                aggregate_week_2=None,
                aggregate_total=None,
            )
            with self.assertRaisesRegex(DataValidationError, "duplicate municipality rows"):
                load_weekly_case_workbook(
                    path,
                    "lyme_cases",
                    WEEKLY_CONFIG,
                    self.municipalities,
                    code_width=3,
                )


if __name__ == "__main__":
    unittest.main()
