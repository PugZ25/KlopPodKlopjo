from __future__ import annotations

import argparse
import csv
import hashlib
import importlib.util
import json
import math
import re
import unicodedata
from collections import Counter
from datetime import date
from itertools import product
from pathlib import Path
from typing import Any, Mapping, Sequence

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_CONFIG_PATH = REPO_ROOT / "model_v3" / "config" / "canonical_epidemiology.json"


class DataValidationError(ValueError):
    """Raised when a source or canonical dataset violates its contract."""


def normalize_text(value: object) -> str:
    """Normalize Unicode and surrounding whitespace without changing wording."""

    return unicodedata.normalize("NFC", str(value)).strip()


def normalize_name_for_lookup(value: object) -> str:
    """Create a case-insensitive, whitespace-normalized municipality lookup key."""

    return " ".join(normalize_text(value).casefold().split())


def canonical_municipality_code(value: object, *, width: int) -> str:
    """Represent an integer municipality code as a fixed-width string."""

    if isinstance(value, bool):
        raise DataValidationError(f"Municipality code cannot be boolean: {value!r}")
    if isinstance(value, int):
        numeric = value
    elif isinstance(value, str) and value.strip().isdigit():
        numeric = int(value.strip())
    else:
        raise DataValidationError(f"Municipality code is not an integer value: {value!r}")

    if numeric < 0 or numeric >= 10**width:
        raise DataValidationError(
            f"Municipality code {numeric!r} does not fit width {width}."
        )
    return f"{numeric:0{width}d}"


def coerce_nonnegative_integer(value: object, *, context: str) -> int | None:
    """Preserve missing values and validate present values as non-negative integers."""

    if value is None:
        return None
    if isinstance(value, bool):
        raise DataValidationError(f"{context} must not be boolean: {value!r}")
    if isinstance(value, int):
        result = value
    elif isinstance(value, float) and math.isfinite(value) and value.is_integer():
        result = int(value)
    else:
        raise DataValidationError(
            f"{context} must be a non-negative integer or missing: {value!r}"
        )
    if result < 0:
        raise DataValidationError(f"{context} must not be negative: {result}")
    return result


def duplicate_key_records(
    rows: Sequence[Mapping[str, object]], keys: Sequence[str]
) -> list[dict[str, object]]:
    counts = Counter(tuple(row[key] for key in keys) for row in rows)
    return [
        {key: value for key, value in zip(keys, values)}
        for values, count in sorted(counts.items(), key=lambda item: repr(item[0]))
        if count > 1
    ]


def require_unique(
    rows: Sequence[Mapping[str, object]], keys: Sequence[str], *, dataset: str
) -> None:
    duplicates = duplicate_key_records(rows, keys)
    if duplicates:
        raise DataValidationError(
            f"{dataset} has unexpected duplicate keys {list(keys)}: {duplicates[:20]}"
        )


def load_config(path: Path) -> dict[str, Any]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if payload.get("schema_version") != 1:
        raise DataValidationError("Canonical configuration schema_version must equal 1.")
    for key in ("municipality_code_width", "sources", "outputs"):
        if key not in payload:
            raise DataValidationError(f"Canonical configuration is missing {key!r}.")
    return payload


def resolve_repo_path(raw_path: object) -> Path:
    if not isinstance(raw_path, str) or not raw_path:
        raise DataValidationError(f"Configured path must be a non-empty string: {raw_path!r}")
    relative = Path(raw_path)
    if relative.is_absolute():
        raise DataValidationError(f"Configured path must be repository-relative: {raw_path}")
    resolved = (REPO_ROOT / relative).resolve()
    if not resolved.is_relative_to(REPO_ROOT):
        raise DataValidationError(f"Configured path leaves repository root: {raw_path}")
    return resolved


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def source_file_record(path: Path) -> dict[str, object]:
    return {
        "path": str(path.relative_to(REPO_ROOT)),
        "bytes": path.stat().st_size,
        "sha256": sha256_file(path),
    }


def load_municipality_dimension(
    source_path: Path,
    source_config: Mapping[str, object],
    *,
    code_width: int,
) -> tuple[list[dict[str, object]], dict[str, object]]:
    payload = json.loads(source_path.read_text(encoding="utf-8"))
    if payload.get("type") != "FeatureCollection":
        raise DataValidationError("GURS municipality source must be a GeoJSON FeatureCollection.")
    features = payload.get("features")
    if not isinstance(features, list) or not features:
        raise DataValidationError("GURS municipality source has no features.")

    code_property = source_config.get("code_property")
    name_property = source_config.get("name_property")
    if not isinstance(code_property, str) or not isinstance(name_property, str):
        raise DataValidationError("Municipality source properties must be configured strings.")

    rows: list[dict[str, object]] = []
    for feature_index, feature in enumerate(features):
        if not isinstance(feature, dict):
            raise DataValidationError(f"GURS feature {feature_index} is not an object.")
        properties = feature.get("properties")
        if not isinstance(properties, dict):
            raise DataValidationError(f"GURS feature {feature_index} has no properties object.")
        if code_property not in properties or name_property not in properties:
            raise DataValidationError(
                f"GURS feature {feature_index} is missing {code_property!r} or {name_property!r}."
            )
        name = normalize_text(properties[name_property])
        if not name:
            raise DataValidationError(f"GURS feature {feature_index} has a blank municipality name.")
        rows.append(
            {
                "municipality_code": canonical_municipality_code(
                    properties[code_property], width=code_width
                ),
                "municipality_name": name,
            }
        )

    require_unique(rows, ["municipality_code"], dataset="municipality")
    lookup_counts = Counter(normalize_name_for_lookup(row["municipality_name"]) for row in rows)
    duplicate_normalized_names = sorted(
        name for name, count in lookup_counts.items() if count > 1
    )
    if duplicate_normalized_names:
        raise DataValidationError(
            "GURS municipality names are ambiguous after lookup normalization: "
            + repr(duplicate_normalized_names)
        )

    rows.sort(key=lambda row: str(row["municipality_code"]))
    quality = {
        "source_feature_count": len(features),
        "duplicate_municipality_codes": [],
        "duplicate_normalized_names": [],
        "blank_codes": 0,
        "blank_names": 0,
    }
    return rows, quality


def ordered_jsonstat_categories(
    dimension: Mapping[str, object], *, dimension_id: str
) -> list[str]:
    category = dimension.get("category")
    if not isinstance(category, dict):
        raise DataValidationError(f"JSON-stat2 dimension {dimension_id!r} has no category.")
    index = category.get("index")
    if not isinstance(index, dict) or not all(
        isinstance(key, str) and isinstance(position, int)
        for key, position in index.items()
    ):
        raise DataValidationError(
            f"JSON-stat2 dimension {dimension_id!r} requires an object category index."
        )
    positions = sorted(index.values())
    if positions != list(range(len(index))):
        raise DataValidationError(
            f"JSON-stat2 dimension {dimension_id!r} has invalid category positions."
        )
    return [key for key, _ in sorted(index.items(), key=lambda item: item[1])]


def jsonstat_status_by_index(payload: Mapping[str, object], value_count: int) -> dict[int, object]:
    status = payload.get("status")
    if status is None:
        return {}
    if isinstance(status, dict):
        result: dict[int, object] = {}
        for raw_index, marker in status.items():
            if not isinstance(raw_index, str) or not raw_index.isdigit():
                raise DataValidationError(f"Invalid JSON-stat2 status index: {raw_index!r}")
            index = int(raw_index)
            if index < 0 or index >= value_count:
                raise DataValidationError(f"JSON-stat2 status index is out of range: {index}")
            result[index] = marker
        return result
    if isinstance(status, list) and len(status) == value_count:
        return {index: marker for index, marker in enumerate(status) if marker is not None}
    raise DataValidationError("Unsupported JSON-stat2 status representation.")


def load_population_dataset(
    source_path: Path,
    source_config: Mapping[str, object],
    municipality_rows: Sequence[Mapping[str, object]],
    *,
    code_width: int,
) -> tuple[list[dict[str, object]], dict[str, object]]:
    payload = json.loads(source_path.read_text(encoding="utf-8"))
    if payload.get("class") != "dataset":
        raise DataValidationError("SURS population source must be a JSON-stat2 dataset.")
    dimension_ids = payload.get("id")
    sizes = payload.get("size")
    dimensions = payload.get("dimension")
    values = payload.get("value")
    if not isinstance(dimension_ids, list) or not all(
        isinstance(item, str) for item in dimension_ids
    ):
        raise DataValidationError("SURS JSON-stat2 id must be a list of dimension names.")
    if not isinstance(sizes, list) or not all(isinstance(item, int) for item in sizes):
        raise DataValidationError("SURS JSON-stat2 size must be an integer list.")
    if not isinstance(dimensions, dict) or not isinstance(values, list):
        raise DataValidationError("SURS JSON-stat2 dimensions or values are invalid.")

    required_config_keys = (
        "measure_dimension",
        "measure_code",
        "expected_measure_label",
        "municipality_dimension",
        "year_dimension",
    )
    if any(not isinstance(source_config.get(key), str) for key in required_config_keys):
        raise DataValidationError("Population dimension configuration is incomplete.")
    measure_dimension = str(source_config["measure_dimension"])
    measure_code = str(source_config["measure_code"])
    expected_measure_label = str(source_config["expected_measure_label"])
    municipality_dimension = str(source_config["municipality_dimension"])
    year_dimension = str(source_config["year_dimension"])
    required_dimensions = {measure_dimension, municipality_dimension, year_dimension}
    if not required_dimensions.issubset(dimension_ids):
        raise DataValidationError(
            f"SURS source is missing configured dimensions: {sorted(required_dimensions - set(dimension_ids))}"
        )

    ordered: dict[str, list[str]] = {}
    for dimension_id in dimension_ids:
        dimension = dimensions.get(dimension_id)
        if not isinstance(dimension, dict):
            raise DataValidationError(f"SURS source is missing dimension {dimension_id!r}.")
        ordered[dimension_id] = ordered_jsonstat_categories(
            dimension, dimension_id=dimension_id
        )
    observed_sizes = [len(ordered[dimension_id]) for dimension_id in dimension_ids]
    if observed_sizes != sizes:
        raise DataValidationError(
            f"SURS JSON-stat2 size mismatch: declared={sizes}, observed={observed_sizes}."
        )
    expected_value_count = math.prod(sizes)
    if len(values) != expected_value_count:
        raise DataValidationError(
            f"SURS JSON-stat2 value count mismatch: expected={expected_value_count}, observed={len(values)}."
        )

    measure = dimensions[measure_dimension]
    measure_category = measure.get("category")
    measure_labels = measure_category.get("label") if isinstance(measure_category, dict) else None
    if not isinstance(measure_labels, dict) or measure_labels.get(measure_code) != expected_measure_label:
        raise DataValidationError(
            f"SURS measure {measure_code!r} does not have expected label {expected_measure_label!r}."
        )
    if measure_code not in ordered[measure_dimension]:
        raise DataValidationError(f"SURS source does not contain measure {measure_code!r}.")

    status_by_index = jsonstat_status_by_index(payload, len(values))
    rows: list[dict[str, object]] = []
    missing_population: list[dict[str, object]] = []
    source_status_markers: Counter[str] = Counter()
    explicit_zero_count = 0
    for flat_index, (coordinates, value) in enumerate(
        zip(product(*(ordered[item] for item in dimension_ids)), values)
    ):
        coordinate = dict(zip(dimension_ids, coordinates))
        if coordinate[measure_dimension] != measure_code:
            continue
        code = canonical_municipality_code(
            coordinate[municipality_dimension], width=code_width
        )
        raw_year = coordinate[year_dimension]
        if not re.fullmatch(r"\d{4}", raw_year):
            raise DataValidationError(f"SURS year is not four digits: {raw_year!r}")
        year = int(raw_year)
        population = coerce_nonnegative_integer(
            value, context=f"population[{code}, {year}]"
        )
        marker = status_by_index.get(flat_index)
        if marker is not None:
            source_status_markers[str(marker)] += 1
        if population is None:
            missing_population.append(
                {
                    "municipality_code": code,
                    "year": year,
                    "source_status": marker,
                }
            )
        elif population == 0:
            explicit_zero_count += 1
        rows.append(
            {
                "municipality_code": code,
                "year": year,
                "population": population,
            }
        )

    require_unique(rows, ["municipality_code", "year"], dataset="population")
    municipality_by_code = {
        str(row["municipality_code"]): str(row["municipality_name"])
        for row in municipality_rows
    }
    population_codes = {str(row["municipality_code"]) for row in rows}
    municipality_codes = set(municipality_by_code)

    municipality_dimension_payload = dimensions[municipality_dimension]
    municipality_category = municipality_dimension_payload.get("category")
    source_labels = (
        municipality_category.get("label")
        if isinstance(municipality_category, dict)
        else None
    )
    if not isinstance(source_labels, dict):
        raise DataValidationError("SURS municipality dimension has no category labels.")
    name_discrepancies = []
    for raw_code in ordered[municipality_dimension]:
        code = canonical_municipality_code(raw_code, width=code_width)
        if code in municipality_by_code and source_labels.get(raw_code) != municipality_by_code[code]:
            name_discrepancies.append(
                {
                    "municipality_code": code,
                    "gurs_name": municipality_by_code[code],
                    "surs_name": source_labels.get(raw_code),
                }
            )

    rows.sort(key=lambda row: (str(row["municipality_code"]), int(row["year"])))
    quality = {
        "measure_code": measure_code,
        "measure_label": expected_measure_label,
        "source_dimension_order": dimension_ids,
        "source_dimension_sizes": sizes,
        "source_value_count": len(values),
        "missing_population_count": len(missing_population),
        "missing_population": missing_population,
        "explicit_zero_count": explicit_zero_count,
        "negative_value_count": 0,
        "duplicate_keys": [],
        "source_status_markers": dict(sorted(source_status_markers.items())),
        "population_codes_not_in_municipality": sorted(population_codes - municipality_codes),
        "municipality_codes_not_in_population": sorted(municipality_codes - population_codes),
        "source_name_discrepancies": name_discrepancies,
    }
    return rows, quality


def build_municipality_name_lookup(
    municipality_rows: Sequence[Mapping[str, object]],
    aliases: Mapping[str, object],
    *,
    code_width: int,
) -> tuple[dict[str, str], dict[str, dict[str, str]]]:
    municipality_by_code = {
        str(row["municipality_code"]): str(row["municipality_name"])
        for row in municipality_rows
    }
    lookup = {
        normalize_name_for_lookup(row["municipality_name"]): str(row["municipality_code"])
        for row in municipality_rows
    }
    alias_details: dict[str, dict[str, str]] = {}
    for raw_name, raw_code in aliases.items():
        if not isinstance(raw_name, str):
            raise DataValidationError(f"Municipality alias name is not a string: {raw_name!r}")
        code = canonical_municipality_code(raw_code, width=code_width)
        if code not in municipality_by_code:
            raise DataValidationError(
                f"Municipality alias {raw_name!r} points to unknown code {code}."
            )
        key = normalize_name_for_lookup(raw_name)
        if key in lookup and lookup[key] != code:
            raise DataValidationError(
                f"Municipality alias {raw_name!r} conflicts with canonical name lookup."
            )
        lookup[key] = code
        alias_details[key] = {
            "source_name": normalize_text(raw_name),
            "municipality_code": code,
            "municipality_name": municipality_by_code[code],
        }
    return lookup, alias_details


def load_weekly_case_workbook(
    source_path: Path,
    case_column: str,
    weekly_config: Mapping[str, object],
    municipality_rows: Sequence[Mapping[str, object]],
    *,
    code_width: int,
) -> tuple[list[dict[str, object]], dict[str, object]]:
    required_integer_keys = ("header_row", "municipality_name_column")
    if any(not isinstance(weekly_config.get(key), int) for key in required_integer_keys):
        raise DataValidationError("Weekly workbook row/column configuration must be integer.")
    header_row = int(weekly_config["header_row"])
    name_column = int(weekly_config["municipality_name_column"])
    municipality_header = weekly_config.get("municipality_header")
    total_label = weekly_config.get("total_label")
    pattern_text = weekly_config.get("week_header_pattern")
    blank_case_cell_value = weekly_config.get("blank_case_cell_value")
    blank_case_cell_rule_source = weekly_config.get("blank_case_cell_rule_source")
    aliases = weekly_config.get("municipality_name_alias_to_code", {})
    if not all(isinstance(value, str) for value in (municipality_header, total_label, pattern_text)):
        raise DataValidationError("Weekly workbook string configuration is incomplete.")
    if not isinstance(aliases, dict):
        raise DataValidationError("Weekly municipality aliases must be an object.")
    if blank_case_cell_value != 0:
        raise DataValidationError(
            "The confirmed NIJZ blank_case_cell_value must be configured as 0."
        )
    if not isinstance(blank_case_cell_rule_source, str) or not blank_case_cell_rule_source:
        raise DataValidationError(
            "Weekly blank_case_cell_rule_source must be a non-empty string."
        )
    week_pattern = re.compile(str(pattern_text))

    name_lookup, alias_details = build_municipality_name_lookup(
        municipality_rows, aliases, code_width=code_width
    )
    municipality_by_code = {
        str(row["municipality_code"]): str(row["municipality_name"])
        for row in municipality_rows
    }
    municipality_codes = set(municipality_by_code)

    workbook = load_workbook(source_path, read_only=True, data_only=True)
    canonical_rows: list[dict[str, object]] = []
    unmatched_names: set[str] = set()
    aliases_used: dict[str, dict[str, str]] = {}
    source_codes: set[str] = set()
    source_blank_case_cell_count = 0
    source_explicit_zero_cell_count = 0
    canonical_zero_count = 0
    sheet_quality: dict[str, object] = {}
    municipality_row_total_mismatches: list[dict[str, object]] = []
    aggregate_total_mismatches: list[dict[str, object]] = []

    for worksheet in workbook.worksheets:
        if not re.fullmatch(r"\d{4}", worksheet.title):
            raise DataValidationError(
                f"Workbook {source_path.name} has non-year sheet {worksheet.title!r}."
            )
        sheet_year = int(worksheet.title)
        matrix = list(worksheet.iter_rows(values_only=True))
        if len(matrix) < header_row:
            raise DataValidationError(f"Sheet {worksheet.title} has no configured header row.")
        header = list(matrix[header_row - 1])
        if len(header) < name_column or header[name_column - 1] != municipality_header:
            observed = header[name_column - 1] if len(header) >= name_column else None
            raise DataValidationError(
                f"Sheet {worksheet.title} municipality header mismatch: {observed!r}."
            )

        week_columns: list[tuple[int, str, int, date]] = []
        for column_index, raw_header in enumerate(header):
            if not isinstance(raw_header, str):
                continue
            match = week_pattern.fullmatch(raw_header.strip())
            if not match:
                continue
            token_year = int(match.group("year"))
            week_number = int(match.group("week"))
            if token_year != sheet_year:
                raise DataValidationError(
                    f"Sheet {sheet_year} contains week token {raw_header!r} from another year."
                )
            try:
                issue_week = date.fromisocalendar(token_year, week_number, 1)
            except ValueError as exc:
                raise DataValidationError(
                    f"Invalid ISO week token in sheet {sheet_year}: {raw_header!r}."
                ) from exc
            week_columns.append((column_index, raw_header.strip(), week_number, issue_week))

        week_tokens = [item[1] for item in week_columns]
        duplicate_week_tokens = sorted(
            token for token, count in Counter(week_tokens).items() if count > 1
        )
        if duplicate_week_tokens:
            raise DataValidationError(
                f"Sheet {sheet_year} has duplicate week headers: {duplicate_week_tokens}."
            )
        expected_week_numbers = set(
            range(1, date(sheet_year, 12, 28).isocalendar().week + 1)
        )
        observed_week_numbers = {item[2] for item in week_columns}
        missing_week_numbers = sorted(expected_week_numbers - observed_week_numbers)
        extra_week_numbers = sorted(observed_week_numbers - expected_week_numbers)

        total_columns = [
            index
            for index, value in enumerate(header)
            if normalize_text(value) == total_label
        ]
        if len(total_columns) != 1:
            raise DataValidationError(
                f"Sheet {sheet_year} must contain exactly one {total_label!r} column."
            )
        total_column = total_columns[0]
        schema_columns = {name_column - 1, total_column, *(item[0] for item in week_columns)}
        unexpected_populated_columns = []
        for column_index in range(len(header)):
            if column_index in schema_columns:
                continue
            nonempty_count = sum(
                row[column_index] is not None
                for row in matrix
                if column_index < len(row)
            )
            if nonempty_count:
                unexpected_populated_columns.append(
                    {
                        "column_index_1_based": column_index + 1,
                        "header": header[column_index],
                        "nonempty_cell_count": nonempty_count,
                    }
                )
        if unexpected_populated_columns:
            raise DataValidationError(
                f"Sheet {sheet_year} has unexpected populated columns: "
                f"{unexpected_populated_columns}."
            )

        named_rows: list[tuple[int, tuple[object, ...], str]] = []
        for row_index, raw_row in enumerate(matrix[header_row:], start=header_row + 1):
            raw_name = raw_row[name_column - 1] if len(raw_row) >= name_column else None
            if raw_name in (None, ""):
                continue
            named_rows.append((row_index, raw_row, normalize_text(raw_name)))
        aggregate_rows = [
            item for item in named_rows if normalize_name_for_lookup(item[2]) == normalize_name_for_lookup(total_label)
        ]
        municipality_source_rows = [item for item in named_rows if item not in aggregate_rows]
        if len(aggregate_rows) != 1:
            raise DataValidationError(
                f"Sheet {sheet_year} must contain exactly one {total_label!r} row."
            )

        normalized_source_names = [normalize_name_for_lookup(item[2]) for item in municipality_source_rows]
        duplicate_source_names = sorted(
            name for name, count in Counter(normalized_source_names).items() if count > 1
        )
        if duplicate_source_names:
            raise DataValidationError(
                f"Sheet {sheet_year} has duplicate municipality rows: {duplicate_source_names}."
            )

        sheet_codes: set[str] = set()
        values_by_week: dict[int, list[int]] = {column_index: [] for column_index, *_ in week_columns}
        for row_index, raw_row, source_name in municipality_source_rows:
            lookup_key = normalize_name_for_lookup(source_name)
            code = name_lookup.get(lookup_key)
            if code is None:
                unmatched_names.add(source_name)
                continue
            if code in sheet_codes:
                raise DataValidationError(
                    f"Sheet {sheet_year} maps multiple municipality rows to code {code}."
                )
            sheet_codes.add(code)
            source_codes.add(code)
            if lookup_key in alias_details:
                aliases_used[lookup_key] = alias_details[lookup_key]

            numeric_sum = 0
            for column_index, token, _, issue_week in week_columns:
                raw_value = raw_row[column_index] if column_index < len(raw_row) else None
                if raw_value is None:
                    source_blank_case_cell_count += 1
                    value = 0
                else:
                    value = coerce_nonnegative_integer(
                        raw_value,
                        context=f"{case_column}[{source_name}, {token}]",
                    )
                    if value == 0:
                        source_explicit_zero_cell_count += 1
                numeric_sum += value
                values_by_week[column_index].append(value)
                if value == 0:
                    canonical_zero_count += 1
                canonical_rows.append(
                    {
                        "municipality_code": code,
                        "issue_week": issue_week,
                        case_column: value,
                    }
                )
            raw_total = raw_row[total_column] if total_column < len(raw_row) else None
            total = (
                0
                if raw_total is None
                else coerce_nonnegative_integer(
                    raw_total,
                    context=f"{case_column}[{source_name}, {total_label}]",
                )
            )
            if total != numeric_sum:
                municipality_row_total_mismatches.append(
                    {
                        "sheet": sheet_year,
                        "row": row_index,
                        "municipality_name": source_name,
                        "numeric_week_sum": numeric_sum,
                        "source_total": total,
                    }
                )

        aggregate_row_index, aggregate_row, _ = aggregate_rows[0]
        aggregate_numeric_sum = 0
        for column_index, token, _, _ in week_columns:
            raw_aggregate = (
                aggregate_row[column_index] if column_index < len(aggregate_row) else None
            )
            aggregate = (
                0
                if raw_aggregate is None
                else coerce_nonnegative_integer(
                    raw_aggregate,
                    context=f"{case_column}[{total_label}, {token}]",
                )
            )
            municipality_numeric_sum = sum(values_by_week[column_index])
            if aggregate != municipality_numeric_sum:
                aggregate_total_mismatches.append(
                    {
                        "sheet": sheet_year,
                        "row": aggregate_row_index,
                        "week": token,
                        "municipality_numeric_sum": municipality_numeric_sum,
                        "source_aggregate": aggregate,
                    }
                )
            aggregate_numeric_sum += aggregate
        raw_grand_total = (
            aggregate_row[total_column] if total_column < len(aggregate_row) else None
        )
        grand_total = (
            0
            if raw_grand_total is None
            else coerce_nonnegative_integer(
                raw_grand_total,
                context=f"{case_column}[{total_label}, {total_label}]",
            )
        )
        if grand_total != aggregate_numeric_sum:
            aggregate_total_mismatches.append(
                {
                    "sheet": sheet_year,
                    "row": aggregate_row_index,
                    "numeric_week_sum": aggregate_numeric_sum,
                    "source_grand_total": grand_total,
                }
            )

        sheet_quality[str(sheet_year)] = {
            "source_week_count": len(week_columns),
            "expected_iso_week_count": len(expected_week_numbers),
            "missing_week_numbers": missing_week_numbers,
            "extra_week_numbers": extra_week_numbers,
            "source_municipality_row_count": len(municipality_source_rows),
            "matched_municipality_code_count": len(sheet_codes),
            "municipality_codes_not_in_sheet": sorted(municipality_codes - sheet_codes),
            "unexpected_duplicate_rows": [],
            "unexpected_populated_columns": [],
        }

    workbook.close()
    if unmatched_names:
        raise DataValidationError(
            f"{source_path.name} has unmatched municipality names: {sorted(unmatched_names)}"
        )
    if municipality_row_total_mismatches:
        raise DataValidationError(
            f"{source_path.name} has municipality-row total mismatches: "
            f"{municipality_row_total_mismatches[:20]}"
        )

    require_unique(
        canonical_rows,
        ["municipality_code", "issue_week"],
        dataset=case_column,
    )
    canonical_rows.sort(
        key=lambda row: (str(row["municipality_code"]), row["issue_week"])
    )
    quality = {
        "source_sheets": list(workbook.sheetnames),
        "sheets": sheet_quality,
        "unmatched_municipality_names": [],
        "source_municipality_codes_not_in_dimension": sorted(source_codes - municipality_codes),
        "municipality_codes_not_in_source": sorted(municipality_codes - source_codes),
        "name_aliases_used": [aliases_used[key] for key in sorted(aliases_used)],
        "blank_case_cell_rule": {
            "canonical_value": 0,
            "source": blank_case_cell_rule_source,
        },
        "source_blank_case_cell_count": source_blank_case_cell_count,
        "source_explicit_zero_cell_count": source_explicit_zero_cell_count,
        "canonical_zero_count": canonical_zero_count,
        "missing_case_value_count": 0,
        "negative_value_count": 0,
        "unexpected_duplicate_rows": [],
        "municipality_row_total_mismatches": [],
        "aggregate_total_mismatches": aggregate_total_mismatches,
    }
    return canonical_rows, quality


def combine_weekly_cases(
    case_datasets: Mapping[str, Sequence[Mapping[str, object]]],
    municipality_rows: Sequence[Mapping[str, object]],
) -> tuple[list[dict[str, object]], dict[str, object]]:
    keyed: dict[str, dict[tuple[str, date], object]] = {}
    for column, rows in case_datasets.items():
        keyed[column] = {
            (str(row["municipality_code"]), row["issue_week"]): row[column]
            for row in rows
        }
        if len(keyed[column]) != len(rows):
            raise DataValidationError(f"{column} contains duplicate municipality-week keys.")

    all_keys = set().union(*(values.keys() for values in keyed.values()))
    columns = list(case_datasets)
    combined = [
        {
            "municipality_code": code,
            "issue_week": issue_week,
            **{column: keyed[column].get((code, issue_week)) for column in columns},
        }
        for code, issue_week in sorted(all_keys)
    ]
    require_unique(
        combined,
        ["municipality_code", "issue_week"],
        dataset="weekly_cases",
    )

    municipality_codes = {str(row["municipality_code"]) for row in municipality_rows}
    weekly_codes = {str(row["municipality_code"]) for row in combined}
    issue_weeks = sorted({row["issue_week"] for row in combined})
    expected_keys = set(product(municipality_codes, issue_weeks))
    missing_grid_keys = sorted(expected_keys - all_keys)
    quality = {
        "source_key_differences": {
            column: {
                "keys_missing_from_source_count": len(all_keys - set(keyed[column])),
                "keys_missing_from_source_sample": [
                    {"municipality_code": code, "issue_week": issue_week.isoformat()}
                    for code, issue_week in sorted(all_keys - set(keyed[column]))[:100]
                ],
            }
            for column in columns
        },
        "weekly_codes_not_in_municipality": sorted(weekly_codes - municipality_codes),
        "municipality_codes_not_in_weekly_cases": sorted(municipality_codes - weekly_codes),
        "missing_municipality_week_keys_count": len(missing_grid_keys),
        "missing_municipality_week_keys_sample": [
            {"municipality_code": code, "issue_week": issue_week.isoformat()}
            for code, issue_week in missing_grid_keys[:100]
        ],
        "unexpected_duplicate_rows": [],
    }
    return combined, quality


def build_calendar(
    weekly_rows: Sequence[Mapping[str, object]],
) -> list[dict[str, object]]:
    issue_weeks = sorted({row["issue_week"] for row in weekly_rows})
    rows: list[dict[str, object]] = []
    for issue_week in issue_weeks:
        if not isinstance(issue_week, date):
            raise DataValidationError(f"issue_week is not a date: {issue_week!r}")
        if issue_week.weekday() != 0:
            raise DataValidationError(f"issue_week is not Monday: {issue_week.isoformat()}")
        iso = issue_week.isocalendar()
        rows.append(
            {
                "issue_week": issue_week,
                "year": iso.year,
                "iso_week": iso.week,
            }
        )
    require_unique(rows, ["issue_week"], dataset="calendar")
    return rows


def csv_value(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, date):
        return value.isoformat()
    return value


def write_csv_dataset(
    path: Path,
    rows: Sequence[Mapping[str, object]],
    columns: Sequence[str],
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(columns), lineterminator="\n")
        writer.writeheader()
        for row in rows:
            writer.writerow({column: csv_value(row[column]) for column in columns})


def dataset_summary(
    *,
    path: Path,
    rows: Sequence[Mapping[str, object]],
    columns: Sequence[str],
    primary_key: Sequence[str],
) -> dict[str, object]:
    missing_values = {
        column: sum(row[column] is None for row in rows) for column in columns
    }
    zero_values = {
        column: sum(
            row[column] == 0 and not isinstance(row[column], bool)
            for row in rows
            if row[column] is not None
        )
        for column in columns
    }
    return {
        "path": str(path.relative_to(REPO_ROOT)),
        "row_count": len(rows),
        "columns": list(columns),
        "primary_key": list(primary_key),
        "unexpected_duplicate_rows": duplicate_key_records(rows, primary_key),
        "missing_values": missing_values,
        "zero_values": zero_values,
        "sha256": sha256_file(path),
        "bytes": path.stat().st_size,
    }


def build_canonical_layer(config_path: Path = DEFAULT_CONFIG_PATH) -> dict[str, object]:
    config_path = config_path.resolve()
    if not config_path.is_relative_to(REPO_ROOT):
        raise DataValidationError(
            f"Canonical configuration must be inside the repository: {config_path}"
        )
    config = load_config(config_path)
    code_width = config.get("municipality_code_width")
    if not isinstance(code_width, int) or code_width <= 0:
        raise DataValidationError("municipality_code_width must be a positive integer.")
    sources = config.get("sources")
    outputs = config.get("outputs")
    if not isinstance(sources, dict) or not isinstance(outputs, dict):
        raise DataValidationError("Canonical sources and outputs must be objects.")

    municipality_config = sources.get("municipality")
    population_config = sources.get("population")
    weekly_config = sources.get("weekly_cases")
    if not all(isinstance(value, dict) for value in (municipality_config, population_config, weekly_config)):
        raise DataValidationError("Canonical source configuration is incomplete.")

    municipality_path = resolve_repo_path(municipality_config.get("path"))
    population_path = resolve_repo_path(population_config.get("path"))
    weekly_files = weekly_config.get("files")
    if not isinstance(weekly_files, list) or not weekly_files:
        raise DataValidationError("At least one weekly case source file is required.")
    weekly_file_records = []
    seen_case_columns: set[str] = set()
    for item in weekly_files:
        if not isinstance(item, dict):
            raise DataValidationError("Each weekly case source must be an object.")
        path = resolve_repo_path(item.get("path"))
        column = item.get("canonical_column")
        if not isinstance(column, str) or not column:
            raise DataValidationError("Each weekly case source needs canonical_column.")
        if column in seen_case_columns:
            raise DataValidationError(f"Duplicate weekly canonical column: {column}")
        seen_case_columns.add(column)
        weekly_file_records.append((column, path))

    for source_path in [municipality_path, population_path, *(path for _, path in weekly_file_records)]:
        if not source_path.is_file():
            raise DataValidationError(f"Configured source file does not exist: {source_path}")

    municipality_rows, municipality_quality = load_municipality_dimension(
        municipality_path, municipality_config, code_width=code_width
    )
    population_rows, population_quality = load_population_dataset(
        population_path,
        population_config,
        municipality_rows,
        code_width=code_width,
    )
    case_datasets: dict[str, list[dict[str, object]]] = {}
    weekly_source_quality: dict[str, object] = {}
    for case_column, source_path in weekly_file_records:
        rows, quality = load_weekly_case_workbook(
            source_path,
            case_column,
            weekly_config,
            municipality_rows,
            code_width=code_width,
        )
        case_datasets[case_column] = rows
        weekly_source_quality[case_column] = quality

    required_case_columns = {"lyme_cases", "kme_cases"}
    if set(case_datasets) != required_case_columns:
        raise DataValidationError(
            f"Weekly sources must produce {sorted(required_case_columns)}; observed {sorted(case_datasets)}."
        )
    weekly_rows, weekly_quality = combine_weekly_cases(
        case_datasets, municipality_rows
    )
    calendar_rows = build_calendar(weekly_rows)

    output_directory = resolve_repo_path(outputs.get("directory"))
    output_names = {
        key: outputs.get(key)
        for key in (
            "municipality",
            "population",
            "weekly_cases",
            "calendar",
            "data_quality_summary",
        )
    }
    if not all(isinstance(value, str) and value for value in output_names.values()):
        raise DataValidationError("Canonical output filenames must be non-empty strings.")
    output_paths = {
        key: output_directory / str(filename) for key, filename in output_names.items()
    }
    if any(path.parent != output_directory for path in output_paths.values()):
        raise DataValidationError("Canonical output filenames must not contain subdirectories.")

    dataset_specs = {
        "municipality": (
            municipality_rows,
            ["municipality_code", "municipality_name"],
            ["municipality_code"],
        ),
        "population": (
            population_rows,
            ["municipality_code", "year", "population"],
            ["municipality_code", "year"],
        ),
        "weekly_cases": (
            weekly_rows,
            ["municipality_code", "issue_week", "lyme_cases", "kme_cases"],
            ["municipality_code", "issue_week"],
        ),
        "calendar": (
            calendar_rows,
            ["issue_week", "year", "iso_week"],
            ["issue_week"],
        ),
    }
    for name, (rows, columns, _) in dataset_specs.items():
        write_csv_dataset(output_paths[name], rows, columns)

    input_sources = {
        "municipality": source_file_record(municipality_path),
        "population": source_file_record(population_path),
        **{
            case_column: source_file_record(source_path)
            for case_column, source_path in weekly_file_records
        },
        "config": source_file_record(config_path.resolve()),
        "builder": source_file_record(Path(__file__).resolve()),
        "requirements": source_file_record(
            REPO_ROOT / "model_v3" / "requirements.txt"
        ),
    }
    datasets = {
        name: dataset_summary(
            path=output_paths[name],
            rows=rows,
            columns=columns,
            primary_key=primary_key,
        )
        for name, (rows, columns, primary_key) in dataset_specs.items()
    }
    all_case_values = [
        row[column]
        for row in weekly_rows
        for column in ("lyme_cases", "kme_cases")
        if row[column] is not None
    ]
    aggregate_total_mismatch_count = sum(
        len(quality["aggregate_total_mismatches"])
        for quality in weekly_source_quality.values()
    )
    summary: dict[str, object] = {
        "schema_version": 1,
        "pipeline": "model_v3.data.canonical_epidemiology",
        "status": (
            "pass_with_warnings" if aggregate_total_mismatch_count else "pass"
        ),
        "output_format": {
            "format": "csv",
            "parquet_supported": bool(
                importlib.util.find_spec("pyarrow")
                or importlib.util.find_spec("fastparquet")
            ),
            "reason": "Neither pyarrow nor fastparquet is installed in the project environment; CSV follows the documented population-null and NIJZ blank-case-to-zero rules.",
        },
        "sources": input_sources,
        "datasets": datasets,
        "checks": {
            "municipality_code_unique": True,
            "municipality_week_unique": True,
            "population_key_unique": True,
            "calendar_issue_week_unique": True,
            "negative_population_count": 0,
            "negative_case_count": 0,
            "minimum_present_case_count": min(all_case_values) if all_case_values else None,
            "source_aggregate_total_mismatch_count": aggregate_total_mismatch_count,
        },
        "municipality_source_quality": municipality_quality,
        "population_source_quality": population_quality,
        "weekly_source_quality": weekly_source_quality,
        "weekly_join_quality": weekly_quality,
        "confirmed_semantics": [
            {
                "item": "NIJZ workbook blank weekly case cells",
                "canonical_value": 0,
                "source": weekly_config["blank_case_cell_rule_source"],
            }
        ],
        "unresolved_meanings": [
            {
                "item": "NIJZ workbook external provenance and reporting-status definition",
                "status": "UNKNOWN",
                "handling": "Recorded source paths and SHA-256 hashes; no additional provenance claim made.",
            },
            {
                "item": "SURS JSON-stat status marker '-'",
                "status": "UNKNOWN",
                "handling": "Population nulls and raw marker counts are preserved in this summary without interpretation.",
            },
            {
                "item": "NIJZ SKUPAJ values that differ from the sum of municipality rows",
                "status": "UNKNOWN",
                "handling": "Canonical municipality cells are preserved; discrepancies are reported as source-quality warnings and are not redistributed or imputed.",
            },
        ],
        "targets_created": False,
    }
    summary_path = output_paths["data_quality_summary"]
    summary_path.parent.mkdir(parents=True, exist_ok=True)
    summary_path.write_text(
        json.dumps(summary, indent=2, ensure_ascii=False, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    return summary


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Build the model_v3 canonical epidemiological data layer."
    )
    parser.add_argument(
        "--config",
        type=Path,
        default=DEFAULT_CONFIG_PATH,
        help="Repository-relative or absolute path to the canonical data configuration.",
    )
    return parser


def main() -> int:
    args = build_parser().parse_args()
    config_path = args.config
    if not config_path.is_absolute():
        config_path = (REPO_ROOT / config_path).resolve()
    summary = build_canonical_layer(config_path)
    print("Canonical epidemiological data layer built.")
    for name, dataset in summary["datasets"].items():
        print(f"- {name}: {dataset['row_count']} rows -> {dataset['path']}")
    print(
        "- data_quality_summary: "
        + str(
            resolve_repo_path(
                load_config(config_path)["outputs"]["directory"]
            )
            / load_config(config_path)["outputs"]["data_quality_summary"]
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
