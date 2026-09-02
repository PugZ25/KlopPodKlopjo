from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
from collections import Counter
from pathlib import Path
from typing import Any, Mapping, Sequence

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_CONFIG_PATH = REPO_ROOT / "model_v3" / "config" / "statistical_regions.json"
EXPECTED_HEADERS = (
    "Raven",
    "Šifra kategorije",
    "Desktriptor",
    "Angleški deskriptor",
    "Šifra starša",
)
REGION_CODE_PATTERN = re.compile(r"^\d{2}$")
MUNICIPALITY_CATEGORY_PATTERN = re.compile(r"^(?P<region>\d{2})\.(?P<municipality>\d{3})$")


class StatisticalRegionError(ValueError):
    """Raised when the official SURS region hierarchy is invalid."""


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def resolve_repo_path(path_value: str | Path, repo_root: Path = REPO_ROOT) -> Path:
    path = Path(path_value)
    return path if path.is_absolute() else repo_root / path


def repository_path(path: Path, repo_root: Path) -> str:
    try:
        return str(path.relative_to(repo_root))
    except ValueError:
        return str(path)


def require_hash(path: Path, expected: str, label: str) -> str:
    actual = sha256_file(path)
    if actual != expected:
        raise StatisticalRegionError(
            f"{label} SHA-256 mismatch: expected {expected}, got {actual}"
        )
    return actual


def load_config(path: Path = DEFAULT_CONFIG_PATH) -> dict[str, Any]:
    with path.open(encoding="utf-8") as handle:
        config = json.load(handle)
    if config.get("schema_version") != 1:
        raise StatisticalRegionError("Unsupported statistical-region schema_version")
    return config


def read_canonical_municipalities(path: Path) -> dict[str, str]:
    result: dict[str, str] = {}
    with path.open(encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        required = {"municipality_code", "municipality_name"}
        if reader.fieldnames is None or not required.issubset(reader.fieldnames):
            raise StatisticalRegionError("Canonical municipality schema is invalid")
        for row in reader:
            code = row["municipality_code"]
            if code in result:
                raise StatisticalRegionError(f"Duplicate canonical municipality code: {code}")
            result[code] = row["municipality_name"]
    return result


def parse_surs_hierarchy(
    path: Path,
    *,
    sheet_name: str,
    region_level: int,
    municipality_level: int,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if workbook.sheetnames != [sheet_name]:
        raise StatisticalRegionError(
            f"Unexpected SURS workbook sheets: {workbook.sheetnames}"
        )
    worksheet = workbook[sheet_name]
    header = tuple(cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1)))
    if header != EXPECTED_HEADERS:
        raise StatisticalRegionError(f"Unexpected SURS hierarchy headers: {header}")

    regions: dict[str, dict[str, Any]] = {}
    municipalities: dict[str, dict[str, Any]] = {}
    for level, category, name_sl, name_en, parent in worksheet.iter_rows(
        min_row=2, values_only=True
    ):
        if level == region_level:
            code = str(category)
            if not REGION_CODE_PATTERN.fullmatch(code):
                raise StatisticalRegionError(f"Invalid statistical-region code: {code}")
            if parent is not None:
                raise StatisticalRegionError(f"Region {code} unexpectedly has a parent")
            if code in regions:
                raise StatisticalRegionError(f"Duplicate statistical-region code: {code}")
            regions[code] = {
                "statistical_region_code": code,
                "statistical_region_name": str(name_sl).strip(),
                "statistical_region_name_en": str(name_en).strip(),
            }
        elif level == municipality_level:
            category_text = str(category)
            match = MUNICIPALITY_CATEGORY_PATTERN.fullmatch(category_text)
            if match is None:
                raise StatisticalRegionError(
                    f"Invalid municipality hierarchy category: {category_text}"
                )
            region_code = match.group("region")
            municipality_code = match.group("municipality")
            if str(parent) != region_code:
                raise StatisticalRegionError(
                    f"Municipality {municipality_code} parent contradicts category code"
                )
            if municipality_code in municipalities:
                raise StatisticalRegionError(
                    f"Duplicate municipality hierarchy code: {municipality_code}"
                )
            municipalities[municipality_code] = {
                "municipality_code": municipality_code,
                "statistical_region_code": region_code,
                "surs_municipality_name": str(name_sl).strip(),
                "surs_municipality_name_en": str(name_en).strip(),
            }

    missing_parents = sorted(
        {row["statistical_region_code"] for row in municipalities.values()} - set(regions)
    )
    if missing_parents:
        raise StatisticalRegionError(f"Municipalities reference missing regions: {missing_parents}")
    return (
        [regions[code] for code in sorted(regions)],
        [municipalities[code] for code in sorted(municipalities)],
    )


def write_csv(
    path: Path, columns: Sequence[str], rows: Sequence[Mapping[str, Any]]
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="raise")
        writer.writeheader()
        writer.writerows(rows)


def file_record(path: Path, repo_root: Path) -> dict[str, Any]:
    return {
        "path": repository_path(path, repo_root),
        "bytes": path.stat().st_size,
        "sha256": sha256_file(path),
    }


def run(
    config_path: Path = DEFAULT_CONFIG_PATH, repo_root: Path = REPO_ROOT
) -> dict[str, Any]:
    config = load_config(config_path)
    source_path = resolve_repo_path(config["source"]["path"], repo_root)
    municipality_path = resolve_repo_path(
        config["canonical_municipality"]["path"], repo_root
    )
    source_hash = require_hash(source_path, config["source"]["sha256"], "SURS hierarchy")
    municipality_hash = require_hash(
        municipality_path,
        config["canonical_municipality"]["sha256"],
        "canonical municipality",
    )
    canonical = read_canonical_municipalities(municipality_path)
    regions, hierarchy_municipalities = parse_surs_hierarchy(
        source_path,
        sheet_name=config["source"]["sheet"],
        region_level=int(config["source"]["region_level"]),
        municipality_level=int(config["source"]["municipality_level"]),
    )
    expected_regions = int(config["policy"]["expected_region_count"])
    expected_municipalities = int(config["policy"]["expected_municipality_count"])
    if len(regions) != expected_regions:
        raise StatisticalRegionError(
            f"Expected {expected_regions} regions, found {len(regions)}"
        )
    hierarchy_by_code = {row["municipality_code"]: row for row in hierarchy_municipalities}
    if len(hierarchy_by_code) != expected_municipalities:
        raise StatisticalRegionError(
            f"Expected {expected_municipalities} municipalities, found {len(hierarchy_by_code)}"
        )
    if set(canonical) != set(hierarchy_by_code):
        raise StatisticalRegionError(
            "SURS hierarchy municipality codes do not exactly match canonical codes"
        )

    mapping_rows = [
        {
            "municipality_code": code,
            "statistical_region_code": hierarchy_by_code[code]["statistical_region_code"],
        }
        for code in sorted(canonical)
    ]
    name_differences = [
        {
            "municipality_code": code,
            "canonical_municipality_name": canonical[code],
            "surs_municipality_name": hierarchy_by_code[code]["surs_municipality_name"],
        }
        for code in sorted(canonical)
        if canonical[code] != hierarchy_by_code[code]["surs_municipality_name"]
    ]

    output_directory = resolve_repo_path(config["outputs"]["directory"], repo_root)
    region_path = output_directory / config["outputs"]["statistical_region"]
    mapping_path = output_directory / config["outputs"]["municipality_statistical_region"]
    quality_path = output_directory / config["outputs"]["quality_summary"]
    write_csv(
        region_path,
        ("statistical_region_code", "statistical_region_name", "statistical_region_name_en"),
        regions,
    )
    write_csv(
        mapping_path,
        ("municipality_code", "statistical_region_code"),
        mapping_rows,
    )
    quality = {
        "schema_version": 1,
        "pipeline": "model_v3.data.statistical_regions",
        "source": {
            **config["source"],
            "actual_sha256": source_hash,
        },
        "canonical_municipality": {
            "path": repository_path(municipality_path, repo_root),
            "actual_sha256": municipality_hash,
        },
        "checks": {
            "region_count": len(regions),
            "municipality_count": len(mapping_rows),
            "canonical_code_parity": True,
            "region_codes_unique": True,
            "municipality_codes_unique": True,
            "all_municipality_region_parents_present": True,
            "name_join_used": False,
            "fixed_mapping_all_analysis_years": config["policy"][
                "fixed_mapping_all_analysis_years"
            ],
        },
        "municipalities_per_region": dict(
            sorted(Counter(row["statistical_region_code"] for row in mapping_rows).items())
        ),
        "name_differences": name_differences,
        "name_difference_count": len(name_differences),
        "outputs": {
            "statistical_region": file_record(region_path, repo_root),
            "municipality_statistical_region": file_record(mapping_path, repo_root),
        },
    }
    quality_path.write_text(
        json.dumps(quality, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    quality["quality_summary"] = file_record(quality_path, repo_root)
    return quality


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Build the canonical SURS statistical-region mapping."
    )
    parser.add_argument("--config", type=Path, default=DEFAULT_CONFIG_PATH)
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    quality = run(args.config)
    print(
        "Created statistical-region mapping: "
        f"{quality['checks']['region_count']} regions, "
        f"{quality['checks']['municipality_count']} municipalities"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
